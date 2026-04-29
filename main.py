import os
import json
import io
import re
import tempfile
from pathlib import Path
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse
from openai import AsyncOpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import httpx
import PyPDF2
from docx import Document
import base64
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

load_dotenv()

APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("APP_PORT", "8000"))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "120"))
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))

app = FastAPI(title="TK AI Generator")

SETTINGS_FILE = Path("settings.json")
DEFAULT_SETTINGS = {
    "provider": "openrouter",
    "api_key": "",
    "model": "anthropic/claude-3.5-sonnet",
    "max_tokens": 3000,
    "master_prompt": """Ты инженер, специалист по формированию технологических карт и работ по ТОиР оборудования.

{file_instruction}

Необходимо заполнить:
1. Столбец "Элемент" — основной крупный элемент, входящий в состав узла. Например: Система смазки.
2. Столбец "Подэлемент" — более мелкий элемент, входящий в состав элемента. Например: Картер.

Правила:
• Каждый новый узел, элемент и подэлемент — в отдельной строке по порядку.
• НЕ вноси как "Элемент" или "Подэлемент": гайки, шайбы, винты, шпильки, хомуты, болты, штифты, шпонки.
• Если в столбцах несколько слов — первое слово всегда существительное, остальные после него.
• Элемент и подэлемент — в единственном числе, именительном падеже.
• Слова нельзя сокращать и заменять синонимами.
• Другие столбцы таблицы не удаляй и не изменяй.

ОТВЕТ ДОЛЖЕН БЫТЬ В СТРОГОМ ФОРМАТЕ:

[ТЕКСТ_ОТВЕТ]
Краткое текстовое описание результата для пользователя.
[/ТЕКСТ_ОТВЕТ]

[ТАБЛИЦА]
Элемент|Подэлемент|Наименование операции|Краткое содержание работ|Вид ТОиР|Периодичность|Норма времени, часов|Количество исполнителей|Профессия/Квалификация|Трудоёмкость, человеко/часов|Наименование ТМЦ|Количество ТМЦ|Единицы измерения ТМЦ|Наименование инструменты|Средства индивидуальной защиты|Требования по безопасности
Система смазки|Картер|Осмотр|Визуальный осмотр картера на наличие трещин и подтёков|ТО-1|4320|2.0|1|Слесарь по ремонту автомобилей, 3 разряд|2.0|||||Каска защитная, 1 шт; Очки защитные, 1 шт; Перчатки защитные, 1 пара|Затормозить технику; Выполнять работы при неработающем двигателе
[/ТАБЛИЦА]

ВАЖНО: Каждая строка таблицы — значения через "|". Всего 16 столбцов. Если данных нет — оставьте пусто (||)."""
}

CSV_HEADERS = [
    "Класс", "Подкласс", "Нормализованный код модели",
    "Элемент", "Подэлемент", "Наименование операции",
    "Краткое содержание работ", "Вид ТОиР", "Периодичность",
    "Норма времени, часов", "Количество исполнителей",
    "Профессия/Квалификация", "Трудоёмкость, человеко/часов",
    "Наименование ТМЦ", "Количество ТМЦ", "Единицы измерения ТМЦ",
    "Наименование инструменты", "Средства индивидуальной защиты",
    "Требования по безопасности"
]


def load_settings():
    """Загрузка настроек с приоритетом .env над settings.json"""
    settings = {
        "provider": "openrouter",
        "api_key": os.getenv("OPENROUTER_API_KEY", ""),
        "model": "anthropic/claude-3.5-sonnet",
        "max_tokens": int(os.getenv("DEFAULT_MAX_TOKENS", "3000")),
        "master_prompt": DEFAULT_SETTINGS["master_prompt"]
    }

    if SETTINGS_FILE.exists():
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                file_settings = json.load(f)
                if not os.getenv("OPENROUTER_API_KEY"):
                    settings["api_key"] = file_settings.get("api_key", "")
                settings.update({k: v for k, v in file_settings.items()
                                 if k not in ("api_key", "provider", "model")})
        except Exception as e:
            print(f"⚠️ Ошибка чтения settings.json: {e}")

    return settings


def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def extract_text_from_file(file_path: str) -> str:
    ext = Path(file_path).suffix.lower()
    text = ""
    try:
        if ext == ".pdf":
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() or ""
        elif ext == ".docx":
            doc = Document(file_path)
            text = "\n".join([p.text for p in doc.paragraphs])
        elif ext in (".txt", ".csv"):
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
    except Exception as e:
        text = f"[Ошибка чтения файла: {e}]"
    return text


def parse_ai_table_response(text_response: str) -> list:
    rows = []
    lines = text_response.strip().split("\n")
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 4:
            # Дополняем до 19 колонок
            while len(parts) < len(CSV_HEADERS):
                parts.append("")
            rows.append(parts[:len(CSV_HEADERS)])
    return rows


def create_xlsx(headers: list, rows: list, class_val: str, subclass_val: str, model_code: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Технологическая карта"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name="Arial", size=9)
            if col_idx == 1 and not value:
                cell.value = class_val
            elif col_idx == 2 and not value:
                cell.value = subclass_val
            elif col_idx == 3 and not value:
                cell.value = model_code

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@retry(stop=stop_after_attempt(MAX_RETRIES),
       wait=wait_exponential(multiplier=1, min=2, max=10),
       retry=retry_if_exception_type((httpx.RequestError, httpx.TimeoutException)))
async def call_ai(messages: list, settings: dict):
    provider = settings.get("provider", "openrouter")
    api_key = settings.get("api_key", "")
    model = settings.get("model", "openai/gpt-4o")
    max_tokens = settings.get("max_tokens", 3000)

    if not api_key and provider != "yandex":
        raise HTTPException(status_code=400, detail="API ключ не установлен в настройках или .env")

    timeout = httpx.Timeout(timeout=REQUEST_TIMEOUT, connect=10.0)

    if provider == "yandex":
        iam_token = os.getenv("YANDEX_IAM_TOKEN", "")
        folder_id = os.getenv("YANDEX_FOLDER_ID", "")
        if not iam_token or not folder_id:
            raise HTTPException(status_code=400, detail="Yandex IAM токен или Folder ID не настроены в .env")

        yandex_messages = []
        for m in messages:
            yandex_messages.append({"role": m["role"], "text": m.get("content", "")})

        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(
                "https://llm.api.cloud.yandex.net/foundationModels/v1/completion",
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {iam_token}",
                    "x-folder-id": folder_id
                },
                json={
                    "modelUri": f"gpt://{folder_id}/yandexgpt/latest",
                    "completionOptions": {
                        "stream": False,
                        "temperature": 0.3,
                        "maxTokens": max_tokens
                    },
                    "messages": yandex_messages
                }
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=resp.status_code, detail=f"Yandex GPT error: {resp.text}")
            data = resp.json()
            return data["result"]["alternatives"][0]["message"]["text"]

    elif provider == "openrouter":
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "http://localhost:8000",
                    "X-Title": "TK AI Generator"
                },
                json={
                    "model": model,
                    "messages": messages,
                    "temperature": float(os.getenv("DEFAULT_TEMPERATURE", "0.3")),
                    "max_tokens": max_tokens
                }
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=resp.status_code, detail=f"OpenRouter error: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]

    elif provider == "openai":
        client = AsyncOpenAI(api_key=api_key, timeout=timeout)
        response = await client.chat.completions.create(
            model=model or "gpt-4o",
            messages=messages,
            temperature=0.3,
            max_tokens=max_tokens
        )
        return response.choices[0].message.content

    elif provider == "huggingface":
        hf_token = os.getenv("HF_API_TOKEN", "")
        if not hf_token:
            raise HTTPException(status_code=400, detail="HF API token не настроен в .env")

        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(
                f"https://api-inference.huggingface.co/models/{model}",
                headers={
                    "Authorization": f"Bearer {hf_token}",
                    "Content-Type": "application/json"
                },
                json={
                    "inputs": "\n".join([f"{m['role']}: {m['content']}" for m in messages]),
                    "parameters": {
                        "max_new_tokens": max_tokens,
                        "temperature": 0.3,
                        "return_full_text": False
                    }
                }
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=resp.status_code, detail=f"HF error: {resp.text}")
            data = resp.json()
            return data[0]["generated_text"] if isinstance(data, list) else str(data)

    else:
        raise HTTPException(status_code=400, detail=f"Неизвестный провайдер: {provider}")


@app.get("/", response_class=HTMLResponse)
async def index():
    html_path = Path("teh_card_2.html")
    if html_path.exists():
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Поместите index.html в папку static/</h1>"


@app.get("/api/settings")
async def get_settings():
    return load_settings()


@app.post("/api/settings")
async def update_settings(settings: dict):
    current = load_settings()
    current.update(settings)
    save_settings(current)
    return {"status": "ok"}


@app.post("/api/chat")
async def chat_endpoint(
    message: str = Form(...),
    model_name: str = Form(""),
    equipment_class: str = Form(""),
    subclass: str = Form(""),
    file: Optional[UploadFile] = File(None),
    model: str = Form(None),
    provider: str = Form(None),
):
    settings = load_settings()

    if provider:
        settings["provider"] = provider
    if model:
        settings["model"] = model

    file_text = ""

    if file and file.filename:
        temp_dir = tempfile.gettempdir()
        safe_name = os.path.basename(file.filename)
        temp_path = os.path.join(temp_dir, safe_name)

        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)

        file_text = extract_text_from_file(temp_path)
        try:
            os.remove(temp_path)
        except Exception:
            pass

    if file_text.strip():
        file_instruction = f"""📄 ТЫ ПОЛУЧИЛ ТЕХНИЧЕСКИЙ ПАСПОРТ для модели "{model_name}".
Используй ТОЛЬКО информацию из этого документа для заполнения таблицы.
Если в документе нет данных по какому-то элементу — оставь соответствующие поля пустыми.
НЕ добавляй информацию из общих знаний, если её нет в документе.

СОДЕРЖИМОЕ ДОКУМЕНТА:
---
{file_text[:40000]}
---"""
    else:
        file_instruction = f"""📋 ТЕХНИЧЕСКИЙ ПАСПОРТ НЕ ЗАГРУЖЕН для модели "{model_name}".
Используй свои общие знания по устройству и обслуживанию оборудования данного типа.
Заполни таблицу максимально полно на основе известных тебе данных по модели "{model_name}".
Если точных данных нет — указывай типовые операции и компоненты для этого класса оборудования."""

    base_prompt = settings.get("master_prompt", "")
    if "{file_instruction}" in base_prompt:
        system_content = base_prompt.format(file_instruction=file_instruction)
    else:
        system_content = f"{file_instruction}\n\n{base_prompt}"

    system_content += """
ОТВЕТ ДОЛЖЕН БЫТЬ В СЛЕДУЮЩЕМ ФОРМАТЕ:

[ТЕКСТ_ОТВЕТ]
Тут краткое текстовое описание результата для пользователя.
[/ТЕКСТ_ОТВЕТ]

[ТАБЛИЦА]
Элемент|Подэлемент|Наименование операции|Краткое содержание работ|Вид ТОиР|Периодичность|Норма времени, часов|Количество исполнителей|Профессия/Квалификация|Трудоёмкость, человеко/часов|Наименование ТМЦ|Количество ТМЦ|Единицы измерения ТМЦ|Наименование инструменты|Средства индивидуальной защиты|Требования по безопасности
Система смазки|Картер|Осмотр|Визуальный осмотр картера на наличие трещин и подтёков|ТО-1|4320|2.0|1|Слесарь по ремонту автомобилей, 3 разряд|2.0|||||Каска защитная, 1 шт; Очки защитные, 1 шт; Перчатки защитные, 1 пара|Затормозить технику; Выполнять работы при неработающем двигателе
[/ТАБЛИЦА]

ВАЖНО: Каждая строка таблицы должна содержать значения через символ "|". Всего 16 столбцов.
Если данные для столбца отсутствуют, оставьте его пустым (просто ||)."""

    user_content = f"Модель: {model_name}\nКласс: {equipment_class}\nПодкласс: {subclass}\n\n"
    user_content += f"Запрос: {message}"

    messages = [
        {"role": "system", "content": system_content},
        {"role": "user", "content": user_content}
    ]

    ai_response = await call_ai(messages, settings)

    text_part = ""
    table_part = ""

    text_match = re.search(r"\[ТЕКСТ_ОТВЕТ\](.*?)\[/ТЕКСТ_ОТВЕТ\]", ai_response, re.DOTALL)
    table_match = re.search(r"\[ТАБЛИЦА\](.*?)\[/ТАБЛИЦА\]", ai_response, re.DOTALL)

    if text_match:
        text_part = text_match.group(1).strip()
    else:
        table_start = ai_response.find("[ТАБЛИЦА]")
        text_part = ai_response[:table_start].strip() if table_start != -1 else ai_response

    if table_match:
        table_part = table_match.group(1).strip()
    else:
        table_part = ai_response

    rows = []
    table_lines = table_part.strip().split("\n")
    for line in table_lines:
        line = line.strip()
        if not line:
            continue
        if "Элемент|Подэлемент" in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 2:
            while len(parts) < 16:
                parts.append("")
            rows.append(parts[:16])

    xlsx_data = None
    if rows:
        full_headers = ["Элемент", "Подэлемент", "Наименование операции",
                        "Краткое содержание работ", "Вид ТОиР", "Периодичность",
                        "Норма времени, часов", "Количество исполнителей",
                        "Профессия/Квалификация", "Трудоёмкость, человеко/часов",
                        "Наименование ТМЦ", "Количество ТМЦ", "Единицы измерения ТМЦ",
                        "Наименование инструменты", "Средства индивидуальной защиты",
                        "Требования по безопасности"]
        xlsx_bytes = create_xlsx(full_headers, rows, equipment_class, subclass, model_name)
        xlsx_data = base64.b64encode(xlsx_bytes).decode("utf-8")

    return {
        "text": text_part,
        "table_rows": rows,
        "xlsx_file": xlsx_data,
        "xlsx_filename": f"ТК_{model_name or 'модель'}_{equipment_class}.xlsx"
    }


@app.get("/api/table_template")
async def get_table_template():
    """Возвращает заголовки таблицы для фронтенда (единый источник)"""
    return {
        "headers": CSV_HEADERS,
        "display_headers": CSV_HEADERS[3:]
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=APP_HOST,
        port=APP_PORT,
        reload=os.getenv("RELOAD", "true").lower() == "true"
    )