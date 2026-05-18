"""
Утилиты для создания XLSX-файлов и вызова парсинга PDF.
"""

import io
from typing import List, Dict, Any
from .parsing.pdf_parser import parse_pdf_bytes
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def create_xlsx(headers: List[str], rows: List[List[str]], class_val: str, subclass_val: str, model_code: str) -> bytes:
    """Создает XLSX файл с технологической картой и возвращает его содержимое в виде байтов."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Технологическая карта"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name="Arial", size=9)
            if col_idx == 1 and not value:
                cell.value = class_val
            elif col_idx == 2 and not value:
                cell.value = subclass_val
            elif col_idx == 3 and not value:
                cell.value = model_code

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_pdf_advanced(file_bytes: bytes) -> Dict[str, Any]:
    """
    Синхронная обёртка над parse_pdf_bytes.
    Вызывай из синхронного кода.
    """
    return parse_pdf_bytes(file_bytes)
